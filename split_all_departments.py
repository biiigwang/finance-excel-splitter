#!/usr/bin/env python3
"""
Finance Excel Splitter - Command Line Interface

This script splits an Excel file containing financial data by department.
Each department gets its own Excel file with all sheets filtered to show
only that department's data.

Usage:
    python split_all_departments.py
    python split_all_departments.py -i input.xlsx -o output_dir
    python split_all_departments.py --input input.xlsx --output output_dir
"""

import argparse
import sys
from pathlib import Path
from typing import Tuple

from openpyxl import load_workbook

from core.sheet_analyzer import SheetAnalyzer
from core.department_index import DepartmentIndex
from core.workbook_builder import WorkbookBuilder


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Split Excel file by department',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default paths (auto-detect department column)
  python split_all_departments.py

  # Specify input and output paths
  python split_all_departments.py -i data/input.xlsx -o output/

  # List all available column names
  python split_all_departments.py -l

  # Specify which column to split by
  python split_all_departments.py -s "科室"

  # Using long parameters
  python split_all_departments.py --input data/input.xlsx --output output/ --split-column "科室"
        """
    )

    parser.add_argument(
        '-i', '--input',
        type=str,
        default='data/试验.xlsx',
        help='Input Excel file path (default: data/试验.xlsx)'
    )

    parser.add_argument(
        '-o', '--output',
        type=str,
        default='output',
        help='Output directory path (default: output)'
    )

    parser.add_argument(
        '--keep-empty-sheets',
        action='store_true',
        help='保留空白子表（默认移除）'
    )

    parser.add_argument(
        '-s', '--split-column',
        type=str,
        default=None,
        help='指定拆分列名（默认：自动检测科室列）'
    )

    parser.add_argument(
        '-l', '--list-columns',
        action='store_true',
        help='列出Excel文件中的所有可用列名并退出'
    )

    return parser.parse_args()


def validate_paths(input_path: str, output_dir: str) -> Tuple[Path, Path]:
    """
    Validate input and output paths.

    Args:
        input_path: Path to input Excel file
        output_dir: Path to output directory

    Returns:
        Tuple of (input_path, output_dir) as Path objects

    Raises:
        SystemExit: If validation fails
    """
    input_file = Path(input_path)
    output_path = Path(output_dir)

    # Check if input file exists
    if not input_file.exists():
        print(f"Error: Input file not found: {input_file.absolute()}")
        sys.exit(1)

    # Check if input is a file
    if not input_file.is_file():
        print(f"Error: Input path is not a file: {input_file.absolute()}")
        sys.exit(1)

    # Check file extension
    if input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
        print(f"Error: Input file must be .xlsx or .xlsm format: {input_file.suffix}")
        sys.exit(1)

    # Create output directory if it doesn't exist
    try:
        output_path.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        print(f"Error: Cannot create output directory: {e}")
        sys.exit(1)

    return input_file, output_path


def main():
    """Main entry point for the CLI application."""
    # Parse command line arguments
    args = parse_arguments()

    # If --list-columns is specified, just list columns and exit
    if args.list_columns:
        input_file = Path(args.input)
        if not input_file.exists():
            print(f"Error: Input file not found: {input_file.absolute()}")
            return 1

        print("Loading Excel file...")
        wb = load_workbook(str(input_file), data_only=True)
        analyzer = SheetAnalyzer(wb, split_column=None)
        headers = analyzer.get_all_unique_headers()
        wb.close()

        print("\nAvailable columns:")
        for header in headers:
            print(f"  - {header}")
        return 0

    # Validate paths
    input_file, output_path = validate_paths(args.input, args.output)

    print("=" * 60)
    print("Finance Excel Splitter - 财务数据拆分工具")
    print("=" * 60)
    print(f"\nInput file: {input_file.absolute()}")
    print(f"Output directory: {output_path.absolute()}")
    if args.split_column:
        print(f"Split column: {args.split_column}")
    else:
        print("Split column: Auto-detect (default)")
    print()

    try:
        # Load workbook with data_only=True to get calculated values
        print("Loading Excel file...")
        wb = load_workbook(str(input_file), data_only=True)

        # Analyze all sheets
        print("Analyzing sheets...")
        analyzer = SheetAnalyzer(wb, split_column=args.split_column)
        sheet_structures = analyzer.analyze_all_sheets()

        if not sheet_structures:
            print("Error: No valid sheets found with '科室' column")
            sys.exit(1)

        print(f"Found {len(sheet_structures)} sheets with department data")

        # Build department index once (caches row indices for all departments)
        print("\nBuilding department index...")
        dept_index = DepartmentIndex(wb, sheet_structures)
        dept_index.build_index()

        # Collect all departments from the index
        all_departments = dept_index.get_departments()

        if not all_departments:
            print("Error: No departments found")
            sys.exit(1)

        print(f"Found {len(all_departments)} departments:")
        for dept in sorted(all_departments):
            print(f"  - {dept}")

        # Build workbooks for each department (using index for caching)
        print("\nGenerating department files...")
        # Note: remove_empty_sheets defaults to True, so we invert the flag
        builder = WorkbookBuilder(
            wb, sheet_structures, output_path, dept_index,
            remove_empty_sheets=not args.keep_empty_sheets
        )
        success_count = 0

        for dept in sorted(all_departments):
            try:
                output_file = builder.build_workbook_for_department(dept)
                print(f"  Created: {output_file.name}")
                success_count += 1
            except Exception as e:
                print(f"  Error creating file for {dept}: {e}")

        # Close the original workbook
        wb.close()

        print()
        print("=" * 60)
        print(f"Processing complete!")
        print(f"Successfully created {success_count}/{len(all_departments)} files")
        print(f"Output directory: {output_path.absolute()}")
        print("=" * 60)

        return 0

    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
