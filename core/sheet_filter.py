"""
Sheet filter module.

Filters sheet data by department.
"""

from typing import List, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from .sheet_structure import SheetStructure


class SheetFilter:
    """
    Filters worksheet data to include only rows for a specific department.
    """

    def __init__(self, worksheet: Worksheet, structure: SheetStructure):
        """
        Initialize the filter with a worksheet and its structure.

        Args:
            worksheet: The openpyxl worksheet to filter
            structure: The structure information for this sheet
        """
        self.worksheet = worksheet
        self.structure = structure

    def filter_by_department(self, department: str) -> List[int]:
        """
        Get row indices that belong to the specified department.

        Args:
            department: The department name to filter by

        Returns:
            List of row indices (1-based) that match the department
        """
        if not self.structure.has_data or not self.structure.dept_col:
            return []

        matching_rows = []
        for row_idx in range(
            self.structure.data_start_row, self.worksheet.max_row + 1
        ):
            dept_value = self.worksheet.cell(
                row=row_idx, column=self.structure.dept_col
            ).value

            if dept_value and str(dept_value).strip() == department:
                matching_rows.append(row_idx)

        return matching_rows

    def get_header_row_data(self) -> List[Tuple[int, any]]:
        """
        Get the header row data as a list of (column_index, value) tuples.

        Returns:
            List of tuples containing column index and header value
        """
        if not self.structure.header_row:
            return []

        headers = []
        for col_idx in range(1, self.worksheet.max_column + 1):
            value = self.worksheet.cell(row=self.structure.header_row, column=col_idx).value
            headers.append((col_idx, value))

        return headers

    def get_row_data(self, row_idx: int) -> List[Tuple[int, any]]:
        """
        Get data from a specific row.

        Args:
            row_idx: The row index (1-based)

        Returns:
            List of tuples containing column index and cell value
        """
        row_data = []
        for col_idx in range(1, self.worksheet.max_column + 1):
            value = self.worksheet.cell(row=row_idx, column=col_idx).value
            row_data.append((col_idx, value))

        return row_data
