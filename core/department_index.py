"""
Department index module.

Provides caching of department row indices to avoid repeated data traversal.
"""

from typing import Dict, List, Set, Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .sheet_structure import SheetStructure


class DepartmentIndex:
    """
    Department row index cache for efficient data filtering.

    This class builds an index mapping each department to its row positions
    in each sheet, avoiding the need to traverse data multiple times.
    """

    # Header values to skip when indexing
    SKIP_HEADERS: Set[str] = {'科室', '绩效科室', '序号'}

    def __init__(self, workbook: Workbook, sheet_structures: Dict[str, SheetStructure]):
        """
        Initialize the index with workbook and sheet structures.

        Args:
            workbook: The openpyxl workbook to index
            sheet_structures: Dictionary mapping sheet names to SheetStructure objects
        """
        self.workbook = workbook
        self.sheet_structures = sheet_structures
        # Index structure: {sheet_name: {department: [row_indices]}}
        self._index: Dict[str, Dict[str, List[int]]] = {}
        self._departments: Set[str] = set()
        self._is_built: bool = False

    def build_index(self) -> 'DepartmentIndex':
        """
        Build the complete index by traversing data once.

        Returns:
            Self for method chaining

        Raises:
            ValueError: If workbook or sheet_structures is invalid
        """
        if not self.workbook:
            raise ValueError("Workbook is required")

        if not self.sheet_structures:
            raise ValueError("Sheet structures are required")

        self._index.clear()
        self._departments.clear()

        # Build index for each sheet
        for sheet_name, structure in self.sheet_structures.items():
            if sheet_name not in self.workbook.sheetnames:
                continue

            if not structure.has_data or not structure.dept_col:
                continue

            worksheet = self.workbook[sheet_name]
            sheet_index = self._build_sheet_index(worksheet, structure)
            self._index[sheet_name] = sheet_index

        self._is_built = True
        return self

    def _build_sheet_index(
        self, worksheet: Worksheet, structure: SheetStructure
    ) -> Dict[str, List[int]]:
        """
        Build index for a single sheet.

        Args:
            worksheet: The worksheet to index
            structure: The sheet structure information

        Returns:
            Dictionary mapping department names to row indices
        """
        sheet_index: Dict[str, List[int]] = {}
        dept_col = structure.dept_col
        data_start = structure.data_start_row

        if not data_start:
            data_start = 1

        # Single pass through the data
        for row_idx in range(data_start, worksheet.max_row + 1):
            dept_value = worksheet.cell(row=row_idx, column=dept_col).value

            if not dept_value:
                continue

            dept_name = str(dept_value).strip()

            if not dept_name or self._is_header_or_number(dept_name):
                continue

            if dept_name not in sheet_index:
                sheet_index[dept_name] = []

            sheet_index[dept_name].append(row_idx)
            self._departments.add(dept_name)

        return sheet_index

    def _is_header_or_number(self, value: str) -> bool:
        """
        Check if a value is a header or pure number.

        Args:
            value: The value to check

        Returns:
            True if value should be skipped
        """
        # Skip header names
        if value in self.SKIP_HEADERS:
            return True

        # Skip pure numbers (like row numbers)
        if value.isdigit():
            return True

        return False

    def get_rows(self, department: str, sheet_name: str) -> List[int]:
        """
        Get row indices for a specific department in a specific sheet.

        Args:
            department: The department name
            sheet_name: The sheet name

        Returns:
            List of row indices (1-based), empty list if not found
        """
        if not self._is_built:
            self.build_index()

        sheet_index = self._index.get(sheet_name, {})
        return sheet_index.get(department, []).copy()

    def get_all_rows_for_department(self, department: str) -> Dict[str, List[int]]:
        """
        Get all row indices for a department across all sheets.

        Args:
            department: The department name

        Returns:
            Dictionary mapping sheet names to row indices
        """
        if not self._is_built:
            self.build_index()

        result = {}
        for sheet_name, sheet_index in self._index.items():
            rows = sheet_index.get(department, [])
            if rows:
                result[sheet_name] = rows.copy()

        return result

    def get_departments(self) -> Set[str]:
        """
        Get all indexed departments.

        Returns:
            Set of department names
        """
        if not self._is_built:
            self.build_index()

        return self._departments.copy()

    def get_sorted_departments(self) -> List[str]:
        """
        Get all departments sorted alphabetically.

        Returns:
            Sorted list of department names
        """
        return sorted(list(self.get_departments()))

    def has_sheet_data(self, sheet_name: str) -> bool:
        """
        Check if a sheet has any indexed data.

        Args:
            sheet_name: The sheet name to check

        Returns:
            True if sheet exists and has data
        """
        if not self._is_built:
            self.build_index()

        return sheet_name in self._index and bool(self._index[sheet_name])

    def get_sheet_departments(self, sheet_name: str) -> Set[str]:
        """
        Get departments that have data in a specific sheet.

        Args:
            sheet_name: The sheet name

        Returns:
            Set of department names with data in this sheet
        """
        if not self._is_built:
            self.build_index()

        sheet_index = self._index.get(sheet_name, {})
        return set(sheet_index.keys())

    @property
    def is_built(self) -> bool:
        """Check if index has been built."""
        return self._is_built

    @property
    def indexed_sheet_count(self) -> int:
        """Get number of sheets with indexed data."""
        if not self._is_built:
            self.build_index()

        return len(self._index)