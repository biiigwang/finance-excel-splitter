"""
Workbook builder module.

Builds new workbooks with filtered data for specific departments.
"""

from typing import List, Dict, Optional
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from .sheet_structure import SheetStructure
from .department_index import DepartmentIndex
from .style_utils import copy_cell_style


def apply_unified_style(cell) -> None:
    """
    Apply unified style to a cell.

    Args:
        cell: The openpyxl cell to apply style to
    """
    # Font: Arial 11
    cell.font = Font(name='Arial', size=11)

    # Border: thin border on all sides
    thin_side = Side(style='thin', color='000000')
    cell.border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    # Fill: white background
    cell.fill = PatternFill(
        fill_type='solid',
        fgColor='FFFFFF',
        bgColor='FFFFFF'
    )

    # Alignment: general alignment
    cell.alignment = Alignment(
        horizontal='general',
        vertical='center',
        wrap_text=False
    )


class WorkbookBuilder:
    """
    Builds new workbooks with filtered data for specific departments.
    """

    def __init__(
        self,
        source_workbook: Workbook,
        sheet_structures: Dict[str, SheetStructure],
        output_dir: Path,
        dept_index: Optional[DepartmentIndex] = None,
        remove_empty_sheets: bool = True,
        style_mode: str = "unified"
    ):
        """
        Initialize the builder with source workbook and structures.

        Args:
            source_workbook: The source openpyxl workbook
            sheet_structures: Dictionary mapping sheet names to SheetStructure objects
            output_dir: Directory to save output files
            dept_index: Optional pre-built department index for caching
            remove_empty_sheets: If True, remove sheets with no data for the department
            style_mode: Style mode - "original" (keep original styles),
                       "unified" (apply unified style), or "none" (no styles)
        """
        self.source_workbook = source_workbook
        self.sheet_structures = sheet_structures
        self.output_dir = output_dir
        self.dept_index = dept_index
        self.remove_empty_sheets = remove_empty_sheets
        self.style_mode = style_mode

    def build_workbook_for_department(self, department: str) -> Path:
        """
        Build and save a workbook containing only data for the specified department.

        Args:
            department: The department name to filter by

        Returns:
            Path to the saved file
        """
        new_workbook = self._build_for_department(department)

        # Create safe filename
        safe_name = department.replace('/', '-').replace('\\', '-')
        output_file = self.output_dir / f"{safe_name}.xlsx"

        new_workbook.save(str(output_file))
        new_workbook.close()

        return output_file

    def _build_for_department(self, department: str) -> Workbook:
        """
        Build a new workbook containing only data for the specified department.

        Args:
            department: The department name to filter by

        Returns:
            New openpyxl Workbook with filtered data
        """
        new_workbook = Workbook()
        # Remove the default sheet created by Workbook()
        new_workbook.remove(new_workbook.active)

        for sheet_name in self.source_workbook.sheetnames:
            source_worksheet = self.source_workbook[sheet_name]
            structure = self.sheet_structures.get(sheet_name)

            if not structure or not structure.has_data:
                # Copy sheet as-is if no structure or no department column
                self._copy_sheet_as_is(new_workbook, source_worksheet)
                continue

            # Filter and copy data for this department
            self._copy_filtered_sheet(
                new_workbook, source_worksheet, structure, department
            )

        return new_workbook

    def _copy_sheet_as_is(
        self, target_workbook: Workbook, source_worksheet: Worksheet
    ) -> None:
        """
        Copy a worksheet as-is to the target workbook.

        Args:
            target_workbook: The workbook to copy to
            source_worksheet: The worksheet to copy
        """
        target_worksheet = target_workbook.create_sheet(title=source_worksheet.title)

        for row_idx in range(1, source_worksheet.max_row + 1):
            for col_idx in range(1, source_worksheet.max_column + 1):
                source_cell = source_worksheet.cell(row=row_idx, column=col_idx)
                target_cell = target_worksheet.cell(row=row_idx, column=col_idx)

                target_cell.value = source_cell.value

                # Apply style based on style_mode
                if self.style_mode == "original":
                    copy_cell_style(source_cell, target_cell)
                elif self.style_mode == "unified":
                    apply_unified_style(target_cell)
                # style_mode == "none": do nothing

        # Copy column widths
        for col_letter in source_worksheet.column_dimensions:
            if col_letter in source_worksheet.column_dimensions:
                target_worksheet.column_dimensions[col_letter].width = \
                    source_worksheet.column_dimensions[col_letter].width

        # Copy row heights
        for row_idx in source_worksheet.row_dimensions:
            if row_idx in source_worksheet.row_dimensions:
                target_worksheet.row_dimensions[row_idx].height = \
                    source_worksheet.row_dimensions[row_idx].height

    def _copy_filtered_sheet(
        self,
        target_workbook: Workbook,
        source_worksheet: Worksheet,
        structure: SheetStructure,
        department: str,
    ) -> None:
        """
        Copy a worksheet with filtered data to the target workbook.
        Uses optimized method with index caching if available.

        Args:
            target_workbook: The workbook to copy to
            source_worksheet: The source worksheet
            structure: The sheet structure information
            department: The department to filter by
        """
        # Use optimized method if index is available
        if self.dept_index is not None:
            self._copy_filtered_sheet_optimized(
                target_workbook, source_worksheet, structure, department
            )
        else:
            self._copy_filtered_sheet_legacy(
                target_workbook, source_worksheet, structure, department
            )

    def _copy_filtered_sheet_optimized(
        self,
        target_workbook: Workbook,
        source_worksheet: Worksheet,
        structure: SheetStructure,
        department: str,
    ) -> None:
        """
        Copy a worksheet using optimized method with iter_rows and append.
        Uses cached department index for fast row lookup.

        Args:
            target_workbook: The workbook to copy to
            source_worksheet: The source worksheet
            structure: The sheet structure information
            department: The department to filter by
        """
        # Get matching rows from index (cached)
        matching_rows = self.dept_index.get_rows(department, source_worksheet.title)

        # Check if there is data for this department
        if not matching_rows:
            if self.remove_empty_sheets:
                return  # Skip creating this sheet
            # If keeping empty sheets, create sheet with only header

        target_ws = target_workbook.create_sheet(title=source_worksheet.title)

        # Get total columns
        max_col = source_worksheet.max_column

        # Copy header rows using iter_rows for efficiency
        # Note: iter_rows returns a generator of tuples (one tuple per row)
        data_start = structure.data_start_row or 1
        for row in source_worksheet.iter_rows(min_row=1, max_row=data_start - 1, min_col=1, max_col=max_col):
            row_values = [cell.value for cell in row]
            target_ws.append(row_values)

            # Apply style to header row cells
            if self.style_mode == "original":
                # Get source row for style copying
                source_row_idx = row[0].row
                for col_idx, target_cell in enumerate(target_ws[target_ws.max_row], 1):
                    source_cell = source_worksheet.cell(row=source_row_idx, column=col_idx)
                    copy_cell_style(source_cell, target_cell)
            elif self.style_mode == "unified":
                for cell in target_ws[target_ws.max_row]:
                    apply_unified_style(cell)
            # style_mode == "none": do nothing

        # Copy data rows using iter_rows and append (only if there are matching rows)
        if matching_rows:
            # For each row index, get the row using iter_rows
            for source_row_idx in matching_rows:
                row = next(source_worksheet.iter_rows(min_row=source_row_idx, max_row=source_row_idx, min_col=1, max_col=max_col))
                row_values = [cell.value for cell in row]
                target_ws.append(row_values)

                # Apply style to data row cells
                if self.style_mode == "original":
                    for col_idx, target_cell in enumerate(target_ws[target_ws.max_row], 1):
                        source_cell = source_worksheet.cell(row=source_row_idx, column=col_idx)
                        copy_cell_style(source_cell, target_cell)
                elif self.style_mode == "unified":
                    for cell in target_ws[target_ws.max_row]:
                        apply_unified_style(cell)
                # style_mode == "none": do nothing

        # Copy column widths
        for col_letter in source_worksheet.column_dimensions:
            if source_worksheet.column_dimensions[col_letter].width:
                target_ws.column_dimensions[col_letter].width = \
                    source_worksheet.column_dimensions[col_letter].width

        # Copy row heights for header rows
        for row_idx in range(1, data_start):
            if row_idx in source_worksheet.row_dimensions:
                target_ws.row_dimensions[row_idx].height = \
                    source_worksheet.row_dimensions[row_idx].height

    def _copy_filtered_sheet_legacy(
        self,
        target_workbook: Workbook,
        source_worksheet: Worksheet,
        structure: SheetStructure,
        department: str,
    ) -> None:
        """
        Legacy method for copying filtered sheet without index caching.

        Args:
            target_workbook: The workbook to copy to
            source_worksheet: The source worksheet
            structure: The sheet structure information
            department: The department to filter by
        """
        from .sheet_filter import SheetFilter

        sheet_filter = SheetFilter(source_worksheet, structure)

        # Get matching rows first to check if we should skip this sheet
        matching_rows = sheet_filter.filter_by_department(department)

        # Check if there is data for this department
        if not matching_rows:
            if self.remove_empty_sheets:
                return  # Skip creating this sheet
            # If keeping empty sheets, create sheet with only header

        target_worksheet = target_workbook.create_sheet(title=source_worksheet.title)

        target_row = 1

        # Copy header rows (everything before data starts)
        for row_idx in range(1, structure.data_start_row):
            for col_idx in range(1, source_worksheet.max_column + 1):
                source_cell = source_worksheet.cell(row=row_idx, column=col_idx)
                target_cell = target_worksheet.cell(row=target_row, column=col_idx)

                target_cell.value = source_cell.value

                # Apply style based on style_mode
                if self.style_mode == "original":
                    copy_cell_style(source_cell, target_cell)
                elif self.style_mode == "unified":
                    apply_unified_style(target_cell)
                # style_mode == "none": do nothing

            target_row += 1

        # Copy matching data rows (only if there are matching rows)
        if matching_rows:
            for source_row_idx in matching_rows:
                for col_idx in range(1, source_worksheet.max_column + 1):
                    source_cell = source_worksheet.cell(row=source_row_idx, column=col_idx)
                    target_cell = target_worksheet.cell(row=target_row, column=col_idx)

                    target_cell.value = source_cell.value

                    # Apply style based on style_mode
                    if self.style_mode == "original":
                        copy_cell_style(source_cell, target_cell)
                    elif self.style_mode == "unified":
                        apply_unified_style(target_cell)
                    # style_mode == "none": do nothing

                target_row += 1

        # Copy column widths
        for col_letter in source_worksheet.column_dimensions:
            if col_letter in source_worksheet.column_dimensions:
                target_worksheet.column_dimensions[col_letter].width = \
                    source_worksheet.column_dimensions[col_letter].width

        # Copy row heights for header rows
        for row_idx in range(1, structure.data_start_row):
            if row_idx in source_worksheet.row_dimensions:
                target_worksheet.row_dimensions[row_idx].height = \
                    source_worksheet.row_dimensions[row_idx].height
