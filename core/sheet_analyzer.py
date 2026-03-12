"""
Sheet analyzer module.

Analyzes Excel sheets to find department columns and data structure.
"""

import re
from typing import Optional, List, Dict, Tuple
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .sheet_structure import SheetStructure


class SheetAnalyzer:
    """
    Analyzes Excel sheets to identify department columns and data structure.
    """

    # Pattern to match department column headers (exact match or common variations)
    DEPT_HEADERS = {'科室', '绩效科室', '所属科室', '部门'}

    def __init__(self, workbook: Workbook, split_column: Optional[str] = None):
        """
        Initialize the analyzer with a workbook.

        Args:
            workbook: The openpyxl workbook to analyze
            split_column: Optional name of the column to split by (if None, auto-detect)
        """
        self.workbook = workbook
        self.split_column = split_column

    def analyze_all_sheets(self) -> Dict[str, SheetStructure]:
        """
        Analyze all sheets in the workbook.

        Returns:
            Dictionary mapping sheet names to their SheetStructure
        """
        structures = {}

        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            structure = self._analyze_sheet(worksheet)
            if structure.has_data:
                structures[sheet_name] = structure

        return structures

    def _analyze_sheet(self, worksheet: Worksheet) -> SheetStructure:
        """
        Analyze a single worksheet.

        Args:
            worksheet: The worksheet to analyze

        Returns:
            SheetStructure object containing the analysis results
        """
        structure = SheetStructure(sheet_name=worksheet.title)

        # Find the header row and department column
        header_row, dept_col = self._find_header_and_dept_col(worksheet)

        if header_row and dept_col:
            structure.header_row = header_row
            structure.dept_col = dept_col
            structure.dept_col_letter = self._get_column_letter(dept_col)
            # 检测表头区域的合并单元格，确定实际的数据起始行
            structure.data_start_row = self._find_data_start_row(
                worksheet, header_row
            )
            structure.has_data = True

        return structure

    def _find_data_start_row(
        self, worksheet: Worksheet, header_row: int
    ) -> int:
        """
        根据合并单元格检测实际的数据起始行。

        检查表头行的所有合并单元格，如果有合并单元格跨越了多行，
        则数据起始行应该是这些合并单元格的最大行号 + 1。

        Args:
            worksheet: 工作表
            header_row: 检测到的表头行（包含"科室"列的行）

        Returns:
            数据起始行号（1-based）
        """
        data_start = header_row + 1

        # 检查所有合并单元格
        for merged_range in worksheet.merged_cells.ranges:
            # 如果合并单元格包含表头行的任何部分
            if merged_range.min_row <= header_row <= merged_range.max_row:
                # 数据起始行至少是合并单元格的下一行
                data_start = max(data_start, merged_range.max_row + 1)

        return data_start

    def _find_header_and_dept_col(self, worksheet: Worksheet) -> Tuple[Optional[int], Optional[int]]:
        """
        Find the header row and department column in the sheet.

        Args:
            worksheet: The worksheet to analyze

        Returns:
            Tuple of (header_row, dept_col) or (None, None) if not found
        """
        max_rows_to_check = min(10, worksheet.max_row)

        # If user specified a split column, try to find it first
        if self.split_column:
            target_col_name = self.split_column.strip()
            for row_idx in range(1, max_rows_to_check + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and str(cell_value).strip() == target_col_name:
                        return row_idx, col_idx
            # If specified column not found, return None instead of falling back
            return None, None

        # Auto-detect mode: look for known department headers
        for row_idx in range(1, max_rows_to_check + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value

                # Check for exact match with known department headers
                if cell_value and str(cell_value).strip() in self.DEPT_HEADERS:
                    return row_idx, col_idx

        return None, None

    def _get_column_letter(self, col_idx: int) -> str:
        """
        Convert column index to column letter.

        Args:
            col_idx: 1-based column index

        Returns:
            Column letter (e.g., 'A', 'B', 'AA')
        """
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def get_department_values(self, worksheet: Worksheet, structure: SheetStructure) -> List[str]:
        """
        Get all unique department values from the sheet.

        Args:
            worksheet: The worksheet to read from
            structure: The sheet structure containing department column info

        Returns:
            List of unique department names
        """
        if not structure.has_data or not structure.dept_col:
            return []

        departments = set()
        for row_idx in range(structure.data_start_row, worksheet.max_row + 1):
            dept_value = worksheet.cell(
                row=row_idx, column=structure.dept_col
            ).value

            if dept_value and str(dept_value).strip():
                departments.add(str(dept_value).strip())

        return sorted(list(departments))

    def get_all_unique_headers(self) -> List[str]:
        """
        获取所有 sheet 的唯一列名（表头）。

        对于每个 sheet：
        - 先尝试找到科室列所在的行作为表头行
        - 然后检测所有表头行（考虑合并单元格）
        - 收集所有表头行的非空值

        Returns:
            去重并排序后的列名列表
        """
        all_headers = set()

        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]

            # 先尝试用原逻辑找到表头行
            header_row, _ = self._find_header_row_for_headers(worksheet)

            if header_row:
                # 检测实际的数据起始行（考虑合并单元格）
                data_start = self._find_data_start_row(worksheet, header_row)

                # 收集所有表头行（从第1行到数据起始行的前一行）的非空值
                for row_idx in range(1, data_start):
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            header_str = str(cell_value).strip()
                            if header_str:
                                all_headers.add(header_str)

        # 排序后返回
        return sorted(list(all_headers))

    def _find_header_row_for_headers(self, worksheet: Worksheet) -> Tuple[Optional[int], Optional[int]]:
        """
        仅用于收集列名的内部方法：查找可能的表头行。

        逻辑与 _find_header_and_dept_col 类似，但不依赖 split_column，
        总是尝试自动检测科室列来确定表头行。

        Args:
            worksheet: The worksheet to analyze

        Returns:
            Tuple of (header_row, dept_col) or (None, None) if not found
        """
        max_rows_to_check = min(10, worksheet.max_row)

        # 首先尝试查找科室列
        for row_idx in range(1, max_rows_to_check + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value and str(cell_value).strip() in self.DEPT_HEADERS:
                    return row_idx, col_idx

        # 如果找不到科室列，找列数最多的前几行中的第一行
        max_cols = 0
        best_row = None
        for row_idx in range(1, max_rows_to_check + 1):
            non_empty_count = 0
            for col_idx in range(1, worksheet.max_column + 1):
                if worksheet.cell(row=row_idx, column=col_idx).value:
                    non_empty_count += 1
            if non_empty_count > max_cols:
                max_cols = non_empty_count
                best_row = row_idx

        if best_row:
            return best_row, None

        return None, None
